import pandas as pd
from dotenv import load_dotenv
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook
from ExcelFormatAPI.FormatReportProduction import process_workbook


load_dotenv()

DASH_INVENTORY_HEADERS = [
    'Condition',
    'Item',
    'Brand',
    'Description',
    'SN',
    'Category',
    'Inv-ID',
    'PO-Line',
    'Status'
]


DESKTOP_HEADERS = [
    "PO#",
    "Line#",
    "Condition",
    "MFGR",
    "Item#",
    "Description",
    "IQ Inv-ID",
    "SN",
    "Status",
    "Processor",
    "Type of RAM",
    "RAM Capacity",
    "Drive Caddy",
    "Drive Capacity",
    "Drive Interface",
    "GPU",
    "Wi-Fi",
    "Form Factor",# not in eps
    "Notes",# not in eps
    "Location",
    "Cost",
    "Vendor"# not in eps
]


LAPTOP_HEADERS = [ # = dealt with
    "PO#",
    "Line#",
    "Condition",
    "MFGR",
    "Item#",
    "Description",
    "IQ Inv-ID",
    "SN",
    "Status",
    "Processor",
    "Type of RAM",
    "RAM Capacity",
    "Battery",
    "Touchscreen", # not in eps
    "GPU",
    "Drive Caddy",# not in eps
    "Drive Capacity",
    "Drive Interface",
    "Notes",# not in eps
    "Location",
    "Cost",
    "Vendor"# not in eps
]


EXTREME_LAPTOP_HEADERS = [
    "Item", "Brand", "SN", "Category", "Description", "Condition", "PO-Line", "QTY", "Cost", "Status",
    "Receive Status", "SO-Line", "SO Rep", "Warehouse", "PO Rep", "Update Date", "Customer Asset Tag",
    "Location", "Container ID", "Virtual Container", "Testing Status", "Added", "Date Received", "ID",
    "IQ Inventory ID", "System Manufacturer", "System Model", "System Serial Number",
    "System Serial Number New", "System SKU", "System Processor Information", "System Memory Information",
    "HDD Manufacturer #1", "HDD Model #1", "HDD Serial Number #1", "HDD Capacity #1", "HDD Block Size #1",
    "HDD RPM #1", "HDD Form Factor #1", "HDD Device Type #1", "HDD Health Score #1", "HDD Grade #1",
    "HDD Reallocated Sectors #1", "HDD Power On Hours #1", "HDD Erasure Method #1",
    "HDD Erasure Results #1", "HDD Erasure Start Date #1", "HDD Erasure End Date #1",
    "HDD Erasure Start Time #1", "HDD Erasure End Time #1", "CPU Manufacturer #1", "CPU Model #1",
    "CPU Serial Number #1", "Processor Speed #1", "Processor Cores #1", "Processor Type #1",
    "Memory Manufacturer #1", "Memory Model #1", "Memory Serial Number #1", "Memory Size #1",
    "Memory Speed #1", "Memory Type #1", "Testing Serial Number #1", "Testing Profile #1",
    "Testing CPU #1", "Testing Memory #1", "Testing Battery #1", "Testing Missing Parts #1",
    "Testing Case Condition #1", "Testing Screen Condition #1", "Testing Display Resolution #1",
    "Testing Optical Device #1", "Battery Manufacturer #1", "Battery Model #1", "Battery Serial Number #1",
    "Battery Chemistry #1", "Battery Capacity #1", "Battery Full Charge #1", "Battery Life #1",
    "Battery Drain Test Results #1", "Component Manufacturer #1", "Component Model #1",
    "Component Serial Number #1", "Component Manufacturer #2", "Component Model #2",
    "Component Serial Number #2", "Component Manufacturer #3", "Component Model #3",
    "Component Serial Number #3", "Component Manufacturer #4", "Component Model #4",
    "Component Serial Number #4", "Component Manufacturer #5", "Component Model #5",
    "Component Serial Number #5", "Component Manufacturer #6", "Component Model #6",
    "Component Serial Number #6", "Network Adapter #1", "Video Adapter #1"
]


EXTREME_DESKTOP_HEADERS = [
    "Item", "Brand", "SN", "Category", "Description", "Condition", "PO-Line", "QTY", "Cost", "Status",
    "Receive Status", "SO-Line", "SO Rep", "Warehouse", "PO Rep", "Update Date", "Customer Asset Tag",
    "Location", "Container ID", "Virtual Container", "Testing Status", "Added", "Date Received", "ID",
    "IQ Inventory ID", "System Manufacturer", "System Model", "System Serial Number",
    "System Serial Number New", "System SKU", "System Processor Information", "System Memory Information",
    "HDD Manufacturer #1", "HDD Model #1", "HDD Serial Number #1", "HDD Capacity #1", "HDD Block Size #1",
    "HDD RPM #1", "HDD Form Factor #1", "HDD Device Type #1", "HDD Health Score #1", "HDD Grade #1",
    "HDD Reallocated Sectors #1", "HDD Power On Hours #1", "HDD Erasure Method #1",
    "HDD Erasure Results #1", "HDD Erasure Start Date #1", "HDD Erasure End Date #1",
    "HDD Erasure Start Time #1", "HDD Erasure End Time #1", "CPU Manufacturer #1", "CPU Model #1",
    "CPU Serial Number #1", "Processor Speed #1", "Processor Cores #1", "Processor Type #1",
    "Memory Manufacturer #1", "Memory Model #1", "Memory Serial Number #1", "Memory Size #1",
    "Memory Speed #1", "Memory Type #1", "Testing Serial Number #1", "Testing Profile #1",
    "Testing CPU #1", "Testing Memory #1", "Testing Battery #1", "Testing Missing Parts #1",
    "Testing Case Condition #1", "Testing Screen Condition #1", "Testing Display Resolution #1",
    "Testing Optical Device #1", "Component Manufacturer #1", "Component Model #1",
    "Component Serial Number #1", "Component Manufacturer #2", "Component Model #2",
    "Component Serial Number #2", "Component Manufacturer #3", "Component Model #3",
    "Component Serial Number #3", "Component Manufacturer #4", "Component Model #4",
    "Component Serial Number #4", "Component Manufacturer #5", "Component Model #5",
    "Component Serial Number #5", "Network Adapter #1", "Video Adapter #1"
]


def process_dash(df: pd.DataFrame):
    temp = df.copy()
    temp.rename(columns={'IQ Inventory ID' : 'Inv-ID'}, inplace=True)
    dash_df = temp[DASH_INVENTORY_HEADERS]

    return dash_df


def process_extreme_laptops(df: pd.DataFrame):
    dash_df = process_dash(df)
    # Rename columns to match
    rename = {
        'Item' : 'Item#',
        'System Manufacturer' : 'MFGR',
        'IQ Inventory ID' : 'IQ Inv-ID',
        'System Processor Information' : 'Processor',
        'Memory Type #1' : 'Type of RAM',
        'Memory Size #1' : 'RAM Capacity',
        'Testing Battery #1' : 'Battery',
        'Video Adapter #1' : 'GPU',
        'HDD Capacity #1' : 'Drive Capacity',
        'HDD Form Factor #1' : 'Drive Interface'
    }
    df.rename(columns=rename, inplace=True)

    laptop_df = pd.DataFrame(index=df.index, columns=LAPTOP_HEADERS)

    # split PO line to 2 columns
    po_line_df = df['PO-Line'].str.split('-', expand=True)

    laptop_df['PO#'] = po_line_df[0].values
    laptop_df['Line#'] = po_line_df[1].values

    for col in LAPTOP_HEADERS:
        if col not in df.columns:
            continue
        laptop_df[col] = df[col]


    return dash_df, laptop_df


def process_extreme_desktops(df: pd.DataFrame):
    dash_df = process_dash(df)
    # Rename columns to match
    rename = {
        'System Manufacturer' : 'MFGR',
        'Item' : 'Item#',
        'IQ Inventory ID' : 'IQ Inv-ID',
        'System Processor Information' : 'Processor',
        'Memory Type #1' : 'Type of RAM',
        'System Memory Information' : 'RAM Capacity',
        'HDD Capacity #1' : 'Drive Capacity',
        'HDD Form Factor #1' : 'Drive Interface',
        'Video Adapter #1' : 'GPU',
        'Network Adapter #1' : 'Wi-Fi'
    }
    df.rename(columns=rename, inplace=True)

    desktop_df = pd.DataFrame(index=df.index, columns=DESKTOP_HEADERS)

    # split PO line to 2 columns
    po_line_df = df['PO-Line'].str.split('-', expand=True)

    desktop_df['PO#'] = po_line_df[0].values
    desktop_df['Line#'] = po_line_df[1].values

    for col in DESKTOP_HEADERS:
        if col not in df.columns:
            continue
        desktop_df[col] = df[col]


    return dash_df, desktop_df


def process_extreme_attributes(workbook):
    try:
        raw_dataframe = pd.read_excel(workbook)

        device_type = raw_dataframe['Category'].iloc[0]

        if device_type == 'PC' or device_type == 'Desktop':
            dash, devices = process_extreme_desktops(raw_dataframe)

        elif device_type == 'Laptop':
            dash, devices =  process_extreme_laptops(raw_dataframe)


        else:
            return None

        # Create workbook with raw data
        wb = create_workbook(dash, devices, device_type)

        # Process/format workbook in-memory
        formatted_wb = process_workbook(wb)

        return formatted_wb

    except Exception as ex:
        raise ex

def create_workbook(dash_df, device_df, device_type):
    new_wb = Workbook()
    # Remove default sheet
    default_sheet = new_wb.active
    new_wb.remove(default_sheet)

    # Create first sheet 'Dash Inventory' and append data
    dash_sheet = new_wb.create_sheet(title='Dash Inventory')
    for row in dataframe_to_rows(dash_df, index=False, header=True):
        dash_sheet.append(row)

    # Create second sheet with category name and append data
    type_sheet = new_wb.create_sheet(title=device_type)
    for row in dataframe_to_rows(device_df, index=False, header=True):
        type_sheet.append(row)

    return new_wb


if __name__ == '__main__':
    from ExcelFormatAPI.FormatReportProduction import process_workbook

    # Get dataframes and category string
    dash_data, device_data, category = process_extreme_attributes(os.getenv('AUTO_ATT_TESTING'))

    # Create workbook with raw data
    wb = create_workbook(dash_data, device_data, category)

    # Process/format workbook in-memory
    formatted_wb = process_workbook(wb)

    # Save the formatted workbook
    formatted_wb.save('output.xlsx')
