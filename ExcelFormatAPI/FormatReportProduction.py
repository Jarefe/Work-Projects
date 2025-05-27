import json
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl.worksheet.table import Table
from openpyxl.utils import get_column_letter

# TODO: handle description cleaning and copying to notes

# FORMATTING RULES
# NOTE: can change colors as necessary
# scrap
RED_FILL = PatternFill(bgColor='FFC7CE')

# notes
YELLOW_FILL = PatternFill(bgColor='FFFF00')

# empty values
GRAY_FILL = PatternFill(bgColor='D3D3D3')

# ram
RAM_RULE = {
    'DDR3': 'ADD8E6',  # light blue
    'DDR4': 'C6EFCE',  # light green
    'DDR5': '008000',  # green
}

# style for header
ORANGE_FILL = PatternFill(
    start_color='FFA500',
    end_color='FFA500',
    fill_type='solid'
)

# thin line border
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# left alignment and vertically centered
ALIGNMENT = Alignment(
    horizontal='left',
    vertical='center'
)


def clean_text(text):
    """Removes instances of ' - Dash Specs'"""
    if isinstance(text, str):
        return text.replace(" - Dash Specs", "").strip()
    return text


def create_table(sheet):
    """Creates a table using all the data in the given sheet"""

    # obtain farthest cell
    max_row = sheet.max_row
    max_column = sheet.max_column

    # define range of cells including header
    table_range = f'A1:{get_column_letter(max_column)}{max_row}'

    # create table with data range
    table = Table(displayName=f'Table_{sheet.title.replace(" ", "")}', ref=table_range)
    sheet.add_table(table)


def format_header(sheet):
    """Applies orange fill to header line"""
    for col in range(1, sheet.max_column + 1):
        cell = sheet.cell(row=1, column=col)
        cell.fill = ORANGE_FILL


def autofit(sheet):
    """Loops through each cell to get longest string and apply column spacing accordingly"""
    for col in sheet.columns:
        max_length = 0
        column = col[0].column
        column_letter = get_column_letter(column)

        # get longest cell width in column
        for cell in col:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except Exception:
                pass

        # fit column widths based on max length
        adjusted_width = max_length + 3
        sheet.column_dimensions[column_letter].width = adjusted_width


def check_no_attribute(sheet, check_empty_range):
    """Checks if cell in given range is empty or states 'No Drive'"""
    empty = Rule(
        type='expression',
        formula=['OR(ISBLANK(J2), LOWER(J2)="no drive")'],
        dxf=DifferentialStyle(fill=GRAY_FILL)
    )
    sheet.conditional_formatting.add(check_empty_range, empty)


def apply_conditional_formatting(sheet, sheet_name):
    """Applies conditional formatting to passed in sheet"""
    max_row = sheet.max_row

    # store headers in list to prevent repeated indexing
    headers = [sheet.cell(row=1, column=col).value for col in range(1, sheet.max_column + 1)]

    # create rule specific to dash inventory sheet
    if sheet_name == 'Dash Inventory':
        scrap = Rule(
            type='expression',
            formula=[f'$A2="SCRP"'],
            dxf=DifferentialStyle(fill=RED_FILL)
        )
        sheet.conditional_formatting.add(f'A2:A{max_row}', scrap)
        return

    scrap = Rule(
        type='expression',
        formula=[f'$C2="SCRP"'],
        dxf=DifferentialStyle(fill=RED_FILL)
    )
    sheet.conditional_formatting.add(f'C2:C{max_row}', scrap)

    # memory formatting
    if sheet_name in {'Desktops', 'Laptops', 'Servers'}:
        mem_col = {
            'Desktops': 'K',
            'Laptops': 'K',
            'Servers': 'N'
        }[sheet_name]

        for mem_type, color in RAM_RULE.items():
            rule = Rule(
                type='expression',
                formula=[f'${mem_col}2="{mem_type}"'],
                dxf=DifferentialStyle(fill=PatternFill(bgColor=color))
            )
            sheet.conditional_formatting.add(f'{mem_col}2:{mem_col}{max_row}', rule)

    # look for notes column
    notes_col_index = headers.index('Notes') + 1  # index of notes column
    notes_col_letter = get_column_letter(notes_col_index)

    # highlight non blank cells in notes column
    if notes_col_letter:
        notes_rule = Rule(
            type='expression',
            formula=[f'NOT(ISBLANK(${notes_col_letter}2))'],
            dxf=DifferentialStyle(fill=YELLOW_FILL)
        )
        sheet.conditional_formatting.add(
            f"{notes_col_letter}2:{notes_col_letter}{max_row}", notes_rule
        )

    # apply empty cell formatting
    if sheet_name in ["Desktops", "Laptops"]:
        check_empty_range = f'$J2:$R{max_row}'
    elif sheet_name == "Networking":
        check_empty_range = f'$J2:$O{max_row}'
    elif sheet_name == "Servers":
        check_empty_range = f'$J2:$X{max_row}'
    else:
        check_empty_range = f'$J2:$O{max_row}'  # defaults to smallest range

    check_no_attribute(sheet, check_empty_range)


def FOR_TESTING_convert_to_JSON(workbook):
    """Converts workbook object to JSON and returns JSON data"""

    try:
        json_data = []
        for sheet in workbook.worksheets:
            # create dictionary for each sheet with sheet name and data
            sheet_data = {'sheet_name': sheet.title, 'data': []}
            # get headers from first row of sheet and store in list
            headers = [cell.value for cell in sheet[1]]

            # loop through each row in sheet (skipping headers) and store data in list of dictionaries
            # values_only=True ensures that only values are returned and not cell objects
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # create dictionary for each row with header and value
                row_data = {header: value for header, value in zip(headers, row)}
                sheet_data['data'].append(row_data)

            # append sheet data to json_data list
            json_data.append(sheet_data)

        # convert json_data to JSON string and return
        return json.dumps(json_data)

    except Exception as e:
        print(f"Error converting to JSON: {e}")
        raise

def transform_api_response(api_response):
    """Converts API response to expected format for format_JSON_data"""
    sheets = []

    # Handle top-level key 'data'
    for sheet_name, rows in api_response.get('data', {}).items():
        sheets.append({
            "sheet_name": sheet_name,
            "data": rows
        })

    return sheets


def format_JSON_data(data):
    """Formats passed in JSON data and returns workbook object"""
    try:
        # parse JSON string to Python object if string
        if isinstance(data, str):
            data = json.loads(data)

        data = transform_api_response(data)

        # create new workbook
        wb = Workbook()
        wb.remove(wb.active)

        # create sheets and add data
        for sheet_data in data:
            # create new sheet
            sheet = wb.create_sheet(title=sheet_data['sheet_name'])

            # add headers
            if sheet_data['data']:
                headers = list(sheet_data['data'][0].keys())
                for col, header in enumerate(headers, 1):
                    cell = sheet.cell(row=1, column=col, value=header)
                    cell.border = BORDER
                    cell.alignment = ALIGNMENT

                # add data
                for row_idx, row_data in enumerate(sheet_data['data'], 2):
                    for col_idx, header in enumerate(headers, 1):
                        cell = sheet.cell(row=row_idx, column=col_idx, value=row_data[header])
                        cell.border = BORDER
                        cell.alignment = ALIGNMENT

                # create table
                create_table(sheet)

                # apply orange fill to header row
                format_header(sheet)

                # apply conditional formatting
                apply_conditional_formatting(sheet, sheet_data['sheet_name'])

                # autofit columns
                autofit(sheet)

            else:
                wb.remove(sheet)
                print(f"Deleted empty sheet: {sheet.title}")
                continue

        return wb

    except Exception as e:
        print(f"Error formatting JSON data: {e}")
        raise

from openpyxl.workbook.workbook import Workbook as OpenpyxlWorkbook

def copy_data(old_wb):
    """Copies data from original excel file and inserts into newly created file"""

    # create new workbook
    wb = Workbook()

    # remove default sheet from wb
    wb.remove(wb.active)

    # copy all non empty sheets
    for sheet_name in old_wb.sheetnames:
        original_sheet = old_wb[sheet_name]

        # skip empty sheets
        if is_sheet_empty(original_sheet):
            print(f'{sheet_name} is empty; skipping')
            continue

        # create sheet in new workbook
        new_sheet = wb.create_sheet(title=sheet_name)

        # copy data from old sheet to new sheet columnwise
        for col in original_sheet.iter_cols():
            if col[0].value == "Description":
                for cell in col:
                    cleaned_value = clean_text(cell.value)
                    new_cell = new_sheet.cell(row=cell.row, column=cell.column, value=cleaned_value)
                    new_cell.border = BORDER
                    new_cell.alignment = ALIGNMENT
            else:
                for cell in col:
                    new_cell = new_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
                    new_cell.border = BORDER
                    new_cell.alignment = ALIGNMENT
    return wb

def is_sheet_empty(sheet):
    """Checks if there is information present other than the header line"""
    if sheet.max_row == 1: # if only header is present, then sheet is empty
        return True
    return False


def process_workbook(workbook):
    """Processes workbook object and returns formatted workbook object"""

    # check if valid workbook object was passed in
    try:
        if not isinstance(workbook, OpenpyxlWorkbook):
            raise TypeError("Could not process workbook.")
        print("Valid workbook with sheet names: ", workbook.sheetnames)

        # copy data to new workbook
        wb = copy_data(workbook)

        # go through each sheet in the workbook
        for sheet_name in wb.sheetnames:
            # get current sheet
            current_sheet = wb[sheet_name]

            # create table
            create_table(current_sheet)

            # apply orange fill to header row
            format_header(current_sheet)

            # apply conditional formatting
            apply_conditional_formatting(current_sheet, sheet_name)

            # "autofit" cells
            # can comment out for efficiency
            # has to loop through every single cell in each column to get longest string
            autofit(current_sheet)

        return wb

    except TypeError as e:
        print(f"Error occurred: {e}")
        raise

def format_excel_file(excel_file):
    """Formats passed in excel file and returns workbook object"""
    # parameter will be full excel file object
    try:
        # load workbook from file-like object
        original_wb = load_workbook(excel_file)

        # copy data to new workbook
        wb = copy_data(original_wb)

        # process workbook
        processed_workbook = process_workbook(wb)

        return processed_workbook

    except Exception as e:
        print(f"Error occurred: {e}")
        raise